import asyncio
import os
import sqlite3
from io import BytesIO
from datetime import datetime
from typing import Optional

from aiogram import Bot, Dispatcher, F, Router
from aiogram.client.default import DefaultBotProperties
from aiogram.filters import Command, CommandStart
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import CallbackQuery, Message, ReplyKeyboardMarkup, KeyboardButton, BufferedInputFile
from aiogram.utils.keyboard import InlineKeyboardBuilder
from dotenv import load_dotenv
import psycopg
from psycopg.rows import dict_row
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

load_dotenv()

BOT_TOKEN = os.getenv("BOT_TOKEN")
DATABASE_URL = os.getenv("DATABASE_URL", "").strip()
OWNER_ID_RAW = os.getenv("ADMIN_ID", "").strip()
ADMIN_IDS = {
    int(x.strip()) for x in os.getenv("ADMIN_IDS", "").split(",") if x.strip().isdigit()
}
if OWNER_ID_RAW.isdigit():
    ADMIN_IDS.add(int(OWNER_ID_RAW))
DB_PATH = os.getenv("DB_PATH", "collab_bot.db")
USE_POSTGRES = bool(DATABASE_URL)

router = Router()


# =========================
# Database
# =========================

class DBConnection:
    """Small compatibility wrapper: PostgreSQL on Railway, SQLite fallback for local tests."""

    def __init__(self):
        if USE_POSTGRES:
            self.conn = psycopg.connect(DATABASE_URL, row_factory=dict_row)
        else:
            self.conn = sqlite3.connect(DB_PATH)
            self.conn.row_factory = sqlite3.Row

    def execute(self, sql: str, params=()):
        if USE_POSTGRES:
            sql = sql.replace("?", "%s")
        return self.conn.execute(sql, params)

    def commit(self):
        self.conn.commit()

    def rollback(self):
        self.conn.rollback()

    def close(self):
        self.conn.close()


def db() -> DBConnection:
    return DBConnection()


def init_db() -> None:
    conn = db()
    if USE_POSTGRES:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS users (
                user_id BIGINT PRIMARY KEY,
                username TEXT,
                full_name TEXT,
                role TEXT,
                created_at TEXT,
                referred_by BIGINT,
                referral_created_at TEXT
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS creators (
                user_id BIGINT PRIMARY KEY,
                name TEXT,
                country TEXT,
                region TEXT,
                city TEXT,
                creator_type TEXT,
                niche TEXT,
                work_format TEXT,
                travel_scope TEXT,
                social_link TEXT,
                followers TEXT,
                reach TEXT,
                blog2_platform TEXT,
                blog2_link TEXT,
                blog2_followers TEXT,
                blog2_reach TEXT,
                ad_formats TEXT,
                cooperation_formats TEXT,
                price TEXT,
                excluded_topics TEXT,
                brief_ready TEXT,
                content_types TEXT,
                industries TEXT,
                on_camera TEXT,
                creator_skills TEXT,
                portfolio_link TEXT,
                delivery_available TEXT,
                contact TEXT,
                status TEXT DEFAULT 'active',
                FOREIGN KEY(user_id) REFERENCES users(user_id)
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS business_profiles (
                user_id BIGINT PRIMARY KEY,
                business_name TEXT,
                city TEXT,
                niche TEXT,
                social_link TEXT,
                contact TEXT,
                FOREIGN KEY(user_id) REFERENCES users(user_id)
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS requests (
                request_id BIGSERIAL PRIMARY KEY,
                business_user_id BIGINT,
                business_name TEXT,
                city TEXT,
                niche TEXT,
                task TEXT,
                creator_needed TEXT,
                cooperation_format TEXT,
                budget TEXT,
                social_link TEXT,
                contact TEXT,
                status TEXT DEFAULT 'pending',
                created_at TEXT,
                moderated_at TEXT,
                moderated_by BIGINT,
                FOREIGN KEY(business_user_id) REFERENCES users(user_id)
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS responses (
                response_id BIGSERIAL PRIMARY KEY,
                request_id BIGINT,
                creator_user_id BIGINT,
                business_user_id BIGINT,
                status TEXT DEFAULT 'new',
                created_at TEXT,
                UNIQUE(request_id, creator_user_id),
                FOREIGN KEY(request_id) REFERENCES requests(request_id),
                FOREIGN KEY(creator_user_id) REFERENCES users(user_id),
                FOREIGN KEY(business_user_id) REFERENCES users(user_id)
            )
            """
        )
        # Safe migrations for databases created by an older release.
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS admin_businesses (
                business_id BIGSERIAL PRIMARY KEY,
                business_name TEXT NOT NULL,
                city TEXT,
                niche TEXT,
                social_link TEXT,
                contact TEXT,
                created_at TEXT
            )
            """
        )
        conn.execute("ALTER TABLE requests ADD COLUMN IF NOT EXISTS moderated_at TEXT")
        conn.execute("ALTER TABLE requests ADD COLUMN IF NOT EXISTS moderated_by BIGINT")
        conn.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS referred_by BIGINT")
        conn.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS referral_created_at TEXT")
        for column, coltype in [
            ("country", "TEXT"), ("region", "TEXT"), ("work_format", "TEXT"),
            ("travel_scope", "TEXT"), ("blog2_platform", "TEXT"), ("blog2_link", "TEXT"),
            ("blog2_followers", "TEXT"), ("blog2_reach", "TEXT"), ("ad_formats", "TEXT"),
            ("price", "TEXT"), ("excluded_topics", "TEXT"), ("brief_ready", "TEXT"),
            ("content_types", "TEXT"), ("industries", "TEXT"), ("on_camera", "TEXT"),
            ("creator_skills", "TEXT"), ("portfolio_link", "TEXT"),
            ("delivery_available", "TEXT")
        ]:
            conn.execute(f"ALTER TABLE creators ADD COLUMN IF NOT EXISTS {column} {coltype}")
    else:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS users (
                user_id INTEGER PRIMARY KEY,
                username TEXT,
                full_name TEXT,
                role TEXT,
                created_at TEXT,
                referred_by INTEGER,
                referral_created_at TEXT
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS creators (
                user_id INTEGER PRIMARY KEY,
                name TEXT,
                city TEXT,
                creator_type TEXT,
                niche TEXT,
                social_link TEXT,
                followers TEXT,
                reach TEXT,
                cooperation_formats TEXT,
                contact TEXT,
                status TEXT DEFAULT 'active'
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS business_profiles (
                user_id INTEGER PRIMARY KEY,
                business_name TEXT,
                city TEXT,
                niche TEXT,
                social_link TEXT,
                contact TEXT
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS requests (
                request_id INTEGER PRIMARY KEY AUTOINCREMENT,
                business_user_id INTEGER,
                business_name TEXT,
                city TEXT,
                niche TEXT,
                task TEXT,
                creator_needed TEXT,
                cooperation_format TEXT,
                budget TEXT,
                social_link TEXT,
                contact TEXT,
                status TEXT DEFAULT 'pending',
                created_at TEXT,
                moderated_at TEXT,
                moderated_by INTEGER
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS responses (
                response_id INTEGER PRIMARY KEY AUTOINCREMENT,
                request_id INTEGER,
                creator_user_id INTEGER,
                business_user_id INTEGER,
                status TEXT DEFAULT 'new',
                created_at TEXT,
                UNIQUE(request_id, creator_user_id)
            )
            """
        )
        # SQLite migrations for an existing local DB.
        columns = {r[1] for r in conn.execute("PRAGMA table_info(requests)").fetchall()}
        if "moderated_at" not in columns:
            conn.execute("ALTER TABLE requests ADD COLUMN moderated_at TEXT")
        if "moderated_by" not in columns:
            conn.execute("ALTER TABLE requests ADD COLUMN moderated_by INTEGER")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS admin_businesses (
                business_id INTEGER PRIMARY KEY AUTOINCREMENT,
                business_name TEXT NOT NULL,
                city TEXT,
                niche TEXT,
                social_link TEXT,
                contact TEXT,
                created_at TEXT
            )
            """
        )
        user_columns = {r[1] for r in conn.execute("PRAGMA table_info(users)").fetchall()}
        if "referred_by" not in user_columns:
            conn.execute("ALTER TABLE users ADD COLUMN referred_by INTEGER")
        if "referral_created_at" not in user_columns:
            conn.execute("ALTER TABLE users ADD COLUMN referral_created_at TEXT")
        creator_columns = {r[1] for r in conn.execute("PRAGMA table_info(creators)").fetchall()}
        for column in [
            "country", "region", "work_format", "travel_scope", "blog2_platform",
            "blog2_link", "blog2_followers", "blog2_reach", "ad_formats", "price",
            "excluded_topics", "brief_ready", "content_types", "industries",
            "on_camera", "creator_skills", "portfolio_link", "delivery_available"
        ]:
            if column not in creator_columns:
                conn.execute(f"ALTER TABLE creators ADD COLUMN {column} TEXT")
    conn.commit()
    conn.close()


def upsert_user(message: Message, role: Optional[str] = None, referred_by: Optional[int] = None) -> None:
    """Create/update a user. Referral source is written only on the first visit."""
    conn = db()
    existing = conn.execute("SELECT * FROM users WHERE user_id=?", (message.from_user.id,)).fetchone()
    if existing:
        if role:
            conn.execute(
                "UPDATE users SET role=?, username=?, full_name=? WHERE user_id=?",
                (role, message.from_user.username, message.from_user.full_name, message.from_user.id),
            )
        else:
            conn.execute(
                "UPDATE users SET username=?, full_name=? WHERE user_id=?",
                (message.from_user.username, message.from_user.full_name, message.from_user.id),
            )
    else:
        valid_referrer = None
        if referred_by and referred_by != message.from_user.id:
            inviter = conn.execute("SELECT user_id FROM users WHERE user_id=?", (referred_by,)).fetchone()
            if inviter:
                valid_referrer = referred_by
        conn.execute(
            """INSERT INTO users
            (user_id, username, full_name, role, created_at, referred_by, referral_created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (
                message.from_user.id,
                message.from_user.username,
                message.from_user.full_name,
                role,
                datetime.now().isoformat(),
                valid_referrer,
                datetime.now().isoformat() if valid_referrer else None,
            ),
        )
    conn.commit()
    conn.close()


def parse_referrer(message: Message) -> Optional[int]:
    """Read deep-link payload /start ref_<telegram_id>."""
    text = (message.text or "").strip()
    parts = text.split(maxsplit=1)
    if len(parts) != 2:
        return None
    payload = parts[1].strip()
    if not payload.startswith("ref_"):
        return None
    value = payload[4:]
    return int(value) if value.isdigit() else None


def referral_counts(user_id: int) -> dict:
    conn = db()
    total = conn.execute("SELECT COUNT(*) AS c FROM users WHERE referred_by=?", (user_id,)).fetchone()["c"]
    creators = conn.execute(
        "SELECT COUNT(*) AS c FROM users WHERE referred_by=? AND role='creator'", (user_id,)
    ).fetchone()["c"]
    businesses = conn.execute(
        "SELECT COUNT(*) AS c FROM users WHERE referred_by=? AND role='business'", (user_id,)
    ).fetchone()["c"]
    conn.close()
    return {"total": total, "creators": creators, "businesses": businesses}


def is_admin(user_id: int) -> bool:
    return user_id in ADMIN_IDS

# =========================
# Keyboards
# =========================

def main_menu_kb():
    kb = InlineKeyboardBuilder()
    kb.button(text="Я креатор", callback_data="role_creator_intro")
    kb.button(text="Я бизнес", callback_data="role_business_intro")
    kb.button(text="Как это работает", callback_data="how_it_works")
    kb.adjust(1)
    return kb.as_markup()


def back_to_main_kb():
    kb = InlineKeyboardBuilder()
    kb.button(text="Назад", callback_data="main_menu")
    return kb.as_markup()


def creator_intro_kb():
    kb = InlineKeyboardBuilder()
    kb.button(text="Я блогер", callback_data="creator_type_blogger")
    kb.button(text="Я контент-креатор", callback_data="creator_type_content")
    kb.button(text="Назад", callback_data="main_menu")
    kb.adjust(1)
    return kb.as_markup()


def admin_businesses_kb():
    kb = InlineKeyboardBuilder()
    kb.button(text="➕ Добавить бизнес", callback_data="admin_business_add")
    kb.button(text="📋 Все бизнесы", callback_data="admin_business_list")
    kb.button(text="➕ Создать заявку", callback_data="admin_request_choose_business")
    kb.adjust(1)
    return kb.as_markup()


def admin_business_card_kb(business_id: int):
    kb = InlineKeyboardBuilder()
    kb.button(text="➕ Создать заявку", callback_data=f"admin_request_for:{business_id}")
    kb.button(text="✏️ Редактировать", callback_data=f"admin_business_edit:{business_id}")
    kb.adjust(1)
    return kb.as_markup()


def single_choice_kb(prefix: str, options: list[str]):
    kb = InlineKeyboardBuilder()
    for i, option in enumerate(options):
        kb.button(text=option, callback_data=f"{prefix}:{i}")
    kb.adjust(1)
    return kb.as_markup()


def multi_choice_kb(prefix: str, options: list[str], selected: list[str]):
    kb = InlineKeyboardBuilder()
    for i, option in enumerate(options):
        mark = "✓ " if option in selected else ""
        kb.button(text=f"{mark}{option}", callback_data=f"{prefix}:{i}")
    kb.button(text="Готово", callback_data=f"{prefix}:done")
    kb.adjust(2)
    return kb.as_markup()


BLOG_NICHES = [
    "Lifestyle", "Beauty", "Fashion", "Food", "Travel", "Семья и дети",
    "Спорт", "Бизнес", "Психология", "Образование", "Авто", "Юмор", "Другое"
]
BLOG_PLATFORMS = ["Instagram", "Telegram", "VK", "YouTube", "TikTok", "Threads", "Другое"]
AD_FORMATS = [
    "Stories", "Reels / Shorts", "Посты", "Обзоры / распаковки",
    "Посещение заведений", "Мероприятия", "Амбассадорство", "Другое"
]
COOP_FORMATS = ["Оплата", "Бартер", "Оплата + бартер", "% с продаж", "Амбассадорство"]

CONTENT_TYPES = [
    "UGC-видео", "Reels / Shorts", "Фото", "Обзоры / распаковки",
    "Talking head", "Lifestyle-контент", "Предметная съёмка", "Другое"
]
CONTENT_INDUSTRIES = [
    "Beauty", "Fashion", "Food", "HoReCa", "Travel",
    "Товары", "Эксперты", "Сервисы и приложения", "Спорт", "Другое"
]
CONTENT_SKILLS = [
    "Разработка идеи", "Сценарий", "Съёмка", "Монтаж",
    "Озвучка", "Субтитры", "Полный цикл"
]
CONTENT_COOP_FORMATS = ["Оплата", "Бартер", "Оплата + бартер"]


def business_intro_kb():
    kb = InlineKeyboardBuilder()
    kb.button(text="Создать профиль", callback_data="business_register")
    kb.button(text="Назад", callback_data="main_menu")
    kb.adjust(1)
    return kb.as_markup()


def business_menu_kb():
    kb = InlineKeyboardBuilder()
    kb.button(text="Создать заявку", callback_data="business_create_request")
    kb.button(text="Мои заявки", callback_data="business_my_requests")
    kb.button(text="Отклики креаторов", callback_data="business_responses")
    kb.button(text="Мой профиль", callback_data="business_profile")
    kb.button(text="Изменить профиль", callback_data="business_edit_profile")
    kb.button(text="Пригласить", callback_data="my_referral")
    kb.button(text="Помощь", callback_data="help_business")
    kb.adjust(1)
    return kb.as_markup()


def creator_menu_kb():
    kb = InlineKeyboardBuilder()
    kb.button(text="Проекты", callback_data="creator_view_requests")
    kb.button(text="Мои отклики", callback_data="creator_my_responses")
    kb.button(text="Мой профиль", callback_data="creator_profile")
    kb.button(text="Изменить профиль", callback_data="creator_edit_profile")
    kb.button(text="Пригласить", callback_data="my_referral")
    kb.button(text="Настройки", callback_data="creator_settings")
    kb.button(text="Помощь", callback_data="help_creator")
    kb.adjust(1)
    return kb.as_markup()

def request_kb(request_id: int):
    kb = InlineKeyboardBuilder()
    kb.button(text="Откликнуться", callback_data=f"respond:{request_id}")
    kb.adjust(1)
    return kb.as_markup()


def business_decision_kb(response_id: int):
    kb = InlineKeyboardBuilder()
    kb.button(text="Принять", callback_data=f"accept:{response_id}")
    kb.button(text="Отклонить", callback_data=f"decline:{response_id}")
    kb.adjust(2)
    return kb.as_markup()


def admin_kb():
    kb = InlineKeyboardBuilder()
    kb.button(text="Заявки на модерацию", callback_data="admin_moderation")
    kb.button(text="Статистика", callback_data="admin_stats")
    kb.button(text="Все заявки", callback_data="admin_requests")
    kb.button(text="Выгрузить Excel", callback_data="admin_export")
    kb.adjust(1)
    return kb.as_markup()


def owner_reply_kb():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="🛡 Модерация"), KeyboardButton(text="📊 Статистика")],
            [KeyboardButton(text="👥 Медиа-база"), KeyboardButton(text="🏢 Бизнесы")],
            [KeyboardButton(text="📋 Все заявки"), KeyboardButton(text="💬 Все отклики")],
            [KeyboardButton(text="📥 Выгрузить Excel"), KeyboardButton(text="📣 Рассылка")],
            [KeyboardButton(text="🔗 Рефералы"), KeyboardButton(text="👤 Режим пользователя")],
        ],
        resize_keyboard=True,
        is_persistent=True,
        input_field_placeholder="Панель владельца КЛИК",
    )


def moderation_kb(request_id: int):
    kb = InlineKeyboardBuilder()
    kb.button(text="✅ Одобрить", callback_data=f"moderate_approve:{request_id}")
    kb.button(text="❌ Отклонить", callback_data=f"moderate_reject:{request_id}")
    kb.adjust(2)
    return kb.as_markup()


# =========================
# States
# =========================

class CreatorForm(StatesGroup):
    name = State()
    country = State()
    country_other = State()
    region = State()
    city = State()
    work_format = State()
    travel_scope = State()
    niche = State()
    blog1_platform = State()
    blog1_link = State()
    blog1_followers = State()
    blog1_reach = State()
    add_blog2 = State()
    blog2_platform = State()
    blog2_link = State()
    blog2_followers = State()
    blog2_reach = State()
    ad_formats = State()
    cooperation_formats = State()
    price_choice = State()
    price = State()
    excluded_topics = State()
    brief_ready = State()
    contact = State()


class ContentCreatorForm(StatesGroup):
    name = State()
    country = State()
    country_other = State()
    region = State()
    city = State()
    work_format = State()
    travel_scope = State()
    content_types = State()
    industries = State()
    on_camera = State()
    skills = State()
    portfolio_link = State()
    delivery_available = State()
    cooperation_formats = State()
    price_choice = State()
    price = State()
    excluded_topics = State()
    contact = State()


class AdminBusinessForm(StatesGroup):
    business_name = State()
    city = State()
    niche = State()
    social_link = State()
    contact = State()


class AdminRequestForm(StatesGroup):
    title = State()
    city = State()
    task = State()
    cooperation_format = State()
    budget = State()


class BusinessForm(StatesGroup):
    business_name = State()
    city = State()
    niche = State()
    social_link = State()
    contact = State()


class BroadcastForm(StatesGroup):
    text = State()


class RequestForm(StatesGroup):
    creator_needed = State()
    task = State()
    cooperation_format = State()
    budget = State()


# =========================
# Formatters
# =========================

def request_card(r) -> str:
    return (
        f"<b>Заявка #{r['request_id']}</b>\n\n"
        f"Бизнес: {r['business_name']}\n"
        f"Город: {r['city']}\n"
        f"Ниша: {r['niche']}\n"
        f"Кого ищут: {r['creator_needed']}\n"
        f"Задача: {r['task']}\n"
        f"Формат: {r['cooperation_format']}\n"
        f"Бюджет / бартер: {r['budget']}\n"
        f"Соцсети: {r['social_link']}\n"
        f"Статус: {r['status']}"
    )


def creator_card(c) -> str:
    if (c["creator_type"] or "") == "Контент-креатор":
        return (
            f"<b>Контент-креатор</b>\n\n"
            f"Имя: {c['name']}\n"
            f"География: {c['country'] or '—'}, {c['region'] or '—'}, {c['city'] or '—'}\n"
            f"Формат работы: {c['work_format'] or '—'}\n"
            f"Выезды: {c['travel_scope'] or '—'}\n\n"
            f"Контент: {c['content_types'] or '—'}\n"
            f"Ниши: {c['industries'] or '—'}\n"
            f"В кадре: {c['on_camera'] or '—'}\n"
            f"Навыки: {c['creator_skills'] or '—'}\n"
            f"Портфолио: {c['portfolio_link'] or '—'}\n"
            f"Получение товаров: {c['delivery_available'] or '—'}\n\n"
            f"Сотрудничество: {c['cooperation_formats'] or '—'}\n"
            f"Стоимость: {c['price'] or 'Обсуждается'}\n"
            f"Не работает с: {c['excluded_topics'] or '—'}"
        )

    second_blog = ""
    if c["blog2_link"]:
        second_blog = (
            f"\nБлог 2: {c['blog2_platform'] or '—'} — {c['blog2_link']}\n"
            f"Подписчики: {c['blog2_followers'] or '—'} · Охват: {c['blog2_reach'] or '—'}"
        )
    return (
        f"<b>Блогер</b>\n\n"
        f"Имя: {c['name']}\n"
        f"География: {c['country'] or '—'}, {c['region'] or '—'}, {c['city'] or '—'}\n"
        f"Формат работы: {c['work_format'] or '—'}\n"
        f"Выезды: {c['travel_scope'] or '—'}\n"
        f"Тематика: {c['niche'] or '—'}\n\n"
        f"Блог 1: {c['social_link'] or '—'}\n"
        f"Подписчики: {c['followers'] or '—'} · Охват: {c['reach'] or '—'}"
        f"{second_blog}\n\n"
        f"Рекламные форматы: {c['ad_formats'] or '—'}\n"
        f"Сотрудничество: {c['cooperation_formats'] or '—'}\n"
        f"Стоимость: {c['price'] or 'Обсуждается'}"
    )


def contact_line(user_id: int, contact: str) -> str:
    conn = db()
    u = conn.execute("SELECT username FROM users WHERE user_id=?", (user_id,)).fetchone()
    conn.close()
    tg = f"@{u['username']}" if u and u['username'] else "username не указан"
    return f"Telegram: {tg}\nКонтакт: {contact}"


# =========================
# Handlers
# =========================

@router.message(CommandStart())
async def start(message: Message, state: FSMContext):
    await state.clear()

    # Реферал фиксируется только при первом входе и больше не меняется.
    upsert_user(message, referred_by=parse_referrer(message))

    # Владелец всегда попадает сразу в админ-панель.
    if is_admin(message.from_user.id):
        await message.answer(
            "<b>КЛИК | Панель владельца</b>\n\n"
            "Здесь закреплены модерация, базы, статистика и выгрузка Excel.\n"
            "Для проверки пользовательского сценария нажмите «👤 Режим пользователя».",
            reply_markup=owner_reply_kb(),
        )
        return

    conn = db()
    creator = conn.execute(
        "SELECT user_id FROM creators WHERE user_id=?",
        (message.from_user.id,)
    ).fetchone()
    business = conn.execute(
        "SELECT user_id FROM business_profiles WHERE user_id=?",
        (message.from_user.id,)
    ).fetchone()
    conn.close()

    # Регистрация проходится только один раз.
    if creator:
        await message.answer(
            "<b>КЛИК | Медиа-маркет</b>\n\nРады видеть вас снова.",
            reply_markup=creator_menu_kb()
        )
        return

    if business:
        await message.answer(
            "<b>КЛИК | Медиа-маркет</b>\n\nРады видеть вас снова.",
            reply_markup=business_menu_kb()
        )
        return

    # Первичный экран только для нового пользователя без профиля.
    await message.answer(
        "<b>КЛИК | Медиа-маркет</b>\n\n"
        "Платформа, где <b>бизнес находит креаторов</b>, а <b>креаторы — проекты, рекламу и коллаборации</b>.\n\n"
        "Без бесконечных переписок и поиска по чатам.\n\n"
        "Выберите, кто вы:",
        reply_markup=main_menu_kb()
    )


@router.message(Command("myid"))
async def my_id(message: Message):
    await message.answer(f"Ваш Telegram ID: <code>{message.from_user.id}</code>")


@router.callback_query(F.data == "main_menu")
async def main_menu(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.edit_text(
        "<b>КЛИК | Медиа-маркет</b>\n\nВыберите, кто вы:",
        reply_markup=main_menu_kb()
    )
    await callback.answer()


@router.callback_query(F.data == "how_it_works")
async def how_it_works(callback: CallbackQuery):
    await callback.message.edit_text(
        "<b>Как работает КЛИК</b>\n\n"
        "Бизнес размещает проект и условия сотрудничества.\n\n"
        "Креаторы находят подходящие проекты и нажимают <b>«Откликнуться»</b>.\n\n"
        "Бизнес видит профиль креатора и выбирает: <b>принять</b> или <b>отклонить</b> отклик.\n\n"
        "После принятия КЛИК открывает контакты обеим сторонам.",
        reply_markup=back_to_main_kb()
    )
    await callback.answer()


# ---------- Business registration ----------

@router.callback_query(F.data == "role_business_intro")
async def role_business_intro(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.edit_text(
        "<b>Найдём вам подходящих креаторов.</b>\n\n"
        "Создайте профиль бизнеса, а затем разместите заявку: укажите город, задачу и условия сотрудничества.\n\n"
        "Креаторы смогут откликнуться, а вы — <b>принять или отклонить отклик</b>.",
        reply_markup=business_intro_kb()
    )
    await callback.answer()


@router.callback_query(F.data == "business_register")
async def role_business(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    await state.set_state(BusinessForm.business_name)
    await callback.message.edit_text("<b>Создаём профиль бизнеса</b>\n\nВведите название бизнеса:")
    await callback.answer()


@router.message(BusinessForm.business_name)
async def business_name(message: Message, state: FSMContext):
    await state.update_data(business_name=message.text)
    await state.set_state(BusinessForm.city)
    await message.answer("Город / онлайн:")


@router.message(BusinessForm.city)
async def business_city(message: Message, state: FSMContext):
    await state.update_data(city=message.text)
    await state.set_state(BusinessForm.niche)
    await message.answer("Ниша бизнеса:")


@router.message(BusinessForm.niche)
async def business_niche(message: Message, state: FSMContext):
    await state.update_data(niche=message.text)
    await state.set_state(BusinessForm.social_link)
    await message.answer("Ссылка на соцсети бизнеса:")


@router.message(BusinessForm.social_link)
async def business_social(message: Message, state: FSMContext):
    await state.update_data(social_link=message.text)
    await state.set_state(BusinessForm.contact)
    await message.answer("Контакт для связи:")


@router.message(BusinessForm.contact)
async def business_contact(message: Message, state: FSMContext):
    data = await state.get_data()
    upsert_user(message, "business")
    conn = db()
    conn.execute(
        """
        INSERT INTO business_profiles
        (user_id, business_name, city, niche, social_link, contact)
        VALUES (?, ?, ?, ?, ?, ?)
        ON CONFLICT (user_id) DO UPDATE SET
            business_name=excluded.business_name,
            city=excluded.city,
            niche=excluded.niche,
            social_link=excluded.social_link,
            contact=excluded.contact
        """,
        (message.from_user.id, data["business_name"], data["city"], data["niche"], data["social_link"], message.text)
    )
    conn.commit()
    conn.close()
    await state.clear()
    await message.answer(
        "<b>Готово. Ваш профиль создан.</b>\n\n"
        "Теперь можно разместить первую заявку и начать получать отклики креаторов.",
        reply_markup=business_menu_kb()
    )


# ---------- Creator registration ----------

@router.callback_query(F.data == "role_creator_intro")
async def role_creator_intro(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.edit_text(
        "<b>Кто вы?</b>\n\n"
        "<b>Блогер</b> — у вас есть собственная аудитория, и вы размещаете рекламу и интеграции в своих соцсетях.\n\n"
        "<b>Контент-креатор</b> — вы создаёте фото и видео для брендов. Большая собственная аудитория необязательна.",
        reply_markup=creator_intro_kb()
    )
    await callback.answer()


@router.callback_query(F.data == "creator_type_content")
async def content_creator_start(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    await state.set_state(ContentCreatorForm.name)
    await callback.message.edit_text(
        "<b>Создаём профиль контент-креатора</b>\n\n1. Как вас зовут?"
    )
    await callback.answer()


@router.message(ContentCreatorForm.name)
async def cc_name(message: Message, state: FSMContext):
    await state.update_data(name=message.text.strip())
    await state.set_state(ContentCreatorForm.country)
    await message.answer(
        "2. <b>Где вы находитесь?</b>",
        reply_markup=single_choice_kb("cc_country", ["Россия", "Другая страна"])
    )


@router.callback_query(ContentCreatorForm.country, F.data.startswith("cc_country:"))
async def cc_country(callback: CallbackQuery, state: FSMContext):
    choice = ["Россия", "Другая страна"][int(callback.data.split(":")[1])]
    if choice == "Россия":
        await state.update_data(country="Россия")
        await state.set_state(ContentCreatorForm.region)
        await callback.message.edit_text("3. Укажите регион:")
    else:
        await state.set_state(ContentCreatorForm.country_other)
        await callback.message.edit_text("3. Укажите страну:")
    await callback.answer()


@router.message(ContentCreatorForm.country_other)
async def cc_country_other(message: Message, state: FSMContext):
    await state.update_data(country=message.text.strip())
    await state.set_state(ContentCreatorForm.region)
    await message.answer("Укажите регион / область:")


@router.message(ContentCreatorForm.region)
async def cc_region(message: Message, state: FSMContext):
    await state.update_data(region=message.text.strip())
    await state.set_state(ContentCreatorForm.city)
    await message.answer("4. Укажите город:")


@router.message(ContentCreatorForm.city)
async def cc_city(message: Message, state: FSMContext):
    await state.update_data(city=message.text.strip())
    await state.set_state(ContentCreatorForm.work_format)
    await message.answer(
        "5. <b>В каком формате вы работаете?</b>",
        reply_markup=single_choice_kb("cc_work", ["Онлайн", "Оффлайн", "Онлайн + оффлайн"])
    )


@router.callback_query(ContentCreatorForm.work_format, F.data.startswith("cc_work:"))
async def cc_work(callback: CallbackQuery, state: FSMContext):
    choice = ["Онлайн", "Оффлайн", "Онлайн + оффлайн"][int(callback.data.split(":")[1])]
    await state.update_data(work_format=choice)
    if choice in {"Оффлайн", "Онлайн + оффлайн"}:
        await state.set_state(ContentCreatorForm.travel_scope)
        await callback.message.edit_text(
            "Куда вы готовы выезжать?",
            reply_markup=single_choice_kb("cc_travel", ["Только мой город", "По региону", "По России"])
        )
    else:
        await state.update_data(travel_scope="Не требуется", content_types_selected=[])
        await state.set_state(ContentCreatorForm.content_types)
        await callback.message.edit_text(
            "6. <b>Какой контент вы создаёте?</b>\nМожно выбрать несколько вариантов.",
            reply_markup=multi_choice_kb("cc_content", CONTENT_TYPES, [])
        )
    await callback.answer()


@router.callback_query(ContentCreatorForm.travel_scope, F.data.startswith("cc_travel:"))
async def cc_travel(callback: CallbackQuery, state: FSMContext):
    choice = ["Только мой город", "По региону", "По России"][int(callback.data.split(":")[1])]
    await state.update_data(travel_scope=choice, content_types_selected=[])
    await state.set_state(ContentCreatorForm.content_types)
    await callback.message.edit_text(
        "6. <b>Какой контент вы создаёте?</b>\nМожно выбрать несколько вариантов.",
        reply_markup=multi_choice_kb("cc_content", CONTENT_TYPES, [])
    )
    await callback.answer()


@router.callback_query(ContentCreatorForm.content_types, F.data.startswith("cc_content:"))
async def cc_content_types(callback: CallbackQuery, state: FSMContext):
    value = callback.data.split(":")[1]
    data = await state.get_data()
    selected = list(data.get("content_types_selected", []))
    if value == "done":
        if not selected:
            await callback.answer("Выберите хотя бы один вариант", show_alert=True)
            return
        await state.update_data(content_types=", ".join(selected), industries_selected=[])
        await state.set_state(ContentCreatorForm.industries)
        await callback.message.edit_text(
            "7. <b>С какими нишами работаете?</b>\nМожно выбрать несколько вариантов.",
            reply_markup=multi_choice_kb("cc_industry", CONTENT_INDUSTRIES, [])
        )
        await callback.answer()
        return
    option = CONTENT_TYPES[int(value)]
    selected.remove(option) if option in selected else selected.append(option)
    await state.update_data(content_types_selected=selected)
    await callback.message.edit_reply_markup(reply_markup=multi_choice_kb("cc_content", CONTENT_TYPES, selected))
    await callback.answer()


@router.callback_query(ContentCreatorForm.industries, F.data.startswith("cc_industry:"))
async def cc_industries(callback: CallbackQuery, state: FSMContext):
    value = callback.data.split(":")[1]
    data = await state.get_data()
    selected = list(data.get("industries_selected", []))
    if value == "done":
        if not selected:
            await callback.answer("Выберите хотя бы один вариант", show_alert=True)
            return
        await state.update_data(industries=", ".join(selected))
        await state.set_state(ContentCreatorForm.on_camera)
        await callback.message.edit_text(
            "8. <b>Снимаетесь ли вы сами в кадре?</b>",
            reply_markup=single_choice_kb("cc_camera", ["Да", "Нет", "Могу и в кадре, и за кадром"])
        )
        await callback.answer()
        return
    option = CONTENT_INDUSTRIES[int(value)]
    selected.remove(option) if option in selected else selected.append(option)
    await state.update_data(industries_selected=selected)
    await callback.message.edit_reply_markup(reply_markup=multi_choice_kb("cc_industry", CONTENT_INDUSTRIES, selected))
    await callback.answer()


@router.callback_query(ContentCreatorForm.on_camera, F.data.startswith("cc_camera:"))
async def cc_camera(callback: CallbackQuery, state: FSMContext):
    choice = ["Да", "Нет", "Могу и в кадре, и за кадром"][int(callback.data.split(":")[1])]
    await state.update_data(on_camera=choice, skills_selected=[])
    await state.set_state(ContentCreatorForm.skills)
    await callback.message.edit_text(
        "9. <b>Что вы умеете делать?</b>\nМожно выбрать несколько вариантов.",
        reply_markup=multi_choice_kb("cc_skills", CONTENT_SKILLS, [])
    )
    await callback.answer()


@router.callback_query(ContentCreatorForm.skills, F.data.startswith("cc_skills:"))
async def cc_skills(callback: CallbackQuery, state: FSMContext):
    value = callback.data.split(":")[1]
    data = await state.get_data()
    selected = list(data.get("skills_selected", []))
    if value == "done":
        if not selected:
            await callback.answer("Выберите хотя бы один вариант", show_alert=True)
            return
        await state.update_data(creator_skills=", ".join(selected))
        await state.set_state(ContentCreatorForm.portfolio_link)
        await callback.message.edit_text(
            "10. <b>Добавьте ссылку на портфолио</b>\n\n"
            "Соцсеть, Google Drive, Behance или другая площадка."
        )
        await callback.answer()
        return
    option = CONTENT_SKILLS[int(value)]
    selected.remove(option) if option in selected else selected.append(option)
    await state.update_data(skills_selected=selected)
    await callback.message.edit_reply_markup(reply_markup=multi_choice_kb("cc_skills", CONTENT_SKILLS, selected))
    await callback.answer()


@router.message(ContentCreatorForm.portfolio_link)
async def cc_portfolio(message: Message, state: FSMContext):
    await state.update_data(portfolio_link=message.text.strip())
    await state.set_state(ContentCreatorForm.delivery_available)
    await message.answer(
        "11. <b>Можете получать товары от брендов доставкой?</b>",
        reply_markup=single_choice_kb("cc_delivery", ["Да", "Нет"])
    )


@router.callback_query(ContentCreatorForm.delivery_available, F.data.startswith("cc_delivery:"))
async def cc_delivery(callback: CallbackQuery, state: FSMContext):
    choice = ["Да", "Нет"][int(callback.data.split(":")[1])]
    await state.update_data(delivery_available=choice, cc_coop_selected=[])
    await state.set_state(ContentCreatorForm.cooperation_formats)
    await callback.message.edit_text(
        "12. <b>Какие варианты сотрудничества рассматриваете?</b>\nМожно выбрать несколько вариантов.",
        reply_markup=multi_choice_kb("cc_coop", CONTENT_COOP_FORMATS, [])
    )
    await callback.answer()


@router.callback_query(ContentCreatorForm.cooperation_formats, F.data.startswith("cc_coop:"))
async def cc_coop(callback: CallbackQuery, state: FSMContext):
    value = callback.data.split(":")[1]
    data = await state.get_data()
    selected = list(data.get("cc_coop_selected", []))
    if value == "done":
        if not selected:
            await callback.answer("Выберите хотя бы один вариант", show_alert=True)
            return
        await state.update_data(cooperation_formats=", ".join(selected))
        await state.set_state(ContentCreatorForm.price_choice)
        await callback.message.edit_text(
            "13. <b>Стоимость работы</b>",
            reply_markup=single_choice_kb("cc_price", ["Указать стоимость", "Обсуждается"])
        )
        await callback.answer()
        return
    option = CONTENT_COOP_FORMATS[int(value)]
    selected.remove(option) if option in selected else selected.append(option)
    await state.update_data(cc_coop_selected=selected)
    await callback.message.edit_reply_markup(reply_markup=multi_choice_kb("cc_coop", CONTENT_COOP_FORMATS, selected))
    await callback.answer()


@router.callback_query(ContentCreatorForm.price_choice, F.data.startswith("cc_price:"))
async def cc_price_choice(callback: CallbackQuery, state: FSMContext):
    choice = ["Указать стоимость", "Обсуждается"][int(callback.data.split(":")[1])]
    if choice == "Указать стоимость":
        await state.set_state(ContentCreatorForm.price)
        await callback.message.edit_text("Укажите стоимость работы / стоимость от:")
    else:
        await state.update_data(price="Обсуждается")
        await state.set_state(ContentCreatorForm.excluded_topics)
        await callback.message.edit_text(
            "14. <b>С какими тематиками вы не работаете?</b>\n\n"
            "Напишите ответ или нажмите «Пропустить».",
            reply_markup=single_choice_kb("cc_excluded", ["Пропустить"])
        )
    await callback.answer()


@router.message(ContentCreatorForm.price)
async def cc_price(message: Message, state: FSMContext):
    await state.update_data(price=message.text.strip())
    await state.set_state(ContentCreatorForm.excluded_topics)
    await message.answer(
        "14. <b>С какими тематиками вы не работаете?</b>\n\n"
        "Напишите ответ или нажмите «Пропустить».",
        reply_markup=single_choice_kb("cc_excluded", ["Пропустить"])
    )


@router.callback_query(ContentCreatorForm.excluded_topics, F.data.startswith("cc_excluded:"))
async def cc_excluded_skip(callback: CallbackQuery, state: FSMContext):
    await state.update_data(excluded_topics="Не указано")
    await state.set_state(ContentCreatorForm.contact)
    username = f"@{callback.from_user.username}" if callback.from_user.username else "не указан"
    await callback.message.edit_text(
        "15. <b>Контакт для связи</b>\n\n"
        f"Ваш Telegram: {username}\n"
        "Пришлите дополнительный контакт или напишите «Telegram»."
    )
    await callback.answer()


@router.message(ContentCreatorForm.excluded_topics)
async def cc_excluded(message: Message, state: FSMContext):
    await state.update_data(excluded_topics=message.text.strip())
    await state.set_state(ContentCreatorForm.contact)
    username = f"@{message.from_user.username}" if message.from_user.username else "не указан"
    await message.answer(
        "15. <b>Контакт для связи</b>\n\n"
        f"Ваш Telegram: {username}\n"
        "Пришлите дополнительный контакт или напишите «Telegram»."
    )


@router.message(ContentCreatorForm.contact)
async def cc_contact(message: Message, state: FSMContext):
    data = await state.get_data()
    upsert_user(message, "creator")
    conn = db()
    conn.execute(
        """
        INSERT INTO creators
        (
            user_id, name, country, region, city, creator_type,
            work_format, travel_scope, content_types, industries,
            on_camera, creator_skills, portfolio_link, delivery_available,
            cooperation_formats, price, excluded_topics, contact, status
        )
        VALUES (?, ?, ?, ?, ?, 'Контент-креатор', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'active')
        ON CONFLICT (user_id) DO UPDATE SET
            name=excluded.name,
            country=excluded.country,
            region=excluded.region,
            city=excluded.city,
            creator_type='Контент-креатор',
            work_format=excluded.work_format,
            travel_scope=excluded.travel_scope,
            content_types=excluded.content_types,
            industries=excluded.industries,
            on_camera=excluded.on_camera,
            creator_skills=excluded.creator_skills,
            portfolio_link=excluded.portfolio_link,
            delivery_available=excluded.delivery_available,
            cooperation_formats=excluded.cooperation_formats,
            price=excluded.price,
            excluded_topics=excluded.excluded_topics,
            contact=excluded.contact,
            status='active'
        """,
        (
            message.from_user.id,
            data["name"],
            data.get("country"),
            data.get("region"),
            data.get("city"),
            data.get("work_format"),
            data.get("travel_scope"),
            data.get("content_types"),
            data.get("industries"),
            data.get("on_camera"),
            data.get("creator_skills"),
            data.get("portfolio_link"),
            data.get("delivery_available"),
            data.get("cooperation_formats"),
            data.get("price"),
            data.get("excluded_topics"),
            message.text.strip()
        )
    )
    conn.commit()
    conn.close()
    await state.clear()
    await message.answer(
        "<b>Готово. Профиль контент-креатора создан.</b>\n\n"
        "Теперь можно смотреть актуальные проекты и откликаться на подходящие предложения.",
        reply_markup=creator_menu_kb()
    )


@router.callback_query(F.data.in_({"creator_type_blogger", "creator_register"}))
async def role_creator(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    await state.update_data(creator_type="Блогер")
    await state.set_state(CreatorForm.name)
    await callback.message.edit_text("<b>Создаём профиль блогера</b>\n\n1. Как вас зовут?")
    await callback.answer()


@router.message(CreatorForm.name)
async def creator_name(message: Message, state: FSMContext):
    await state.update_data(name=message.text.strip())
    await state.set_state(CreatorForm.country)
    await message.answer(
        "2. <b>Где вы находитесь?</b>",
        reply_markup=single_choice_kb("blog_country", ["Россия", "Другая страна"])
    )


@router.callback_query(CreatorForm.country, F.data.startswith("blog_country:"))
async def creator_country(callback: CallbackQuery, state: FSMContext):
    idx = int(callback.data.split(":")[1])
    choice = ["Россия", "Другая страна"][idx]
    if choice == "Россия":
        await state.update_data(country="Россия")
        await state.set_state(CreatorForm.region)
        await callback.message.edit_text("3. Укажите ваш регион:")
    else:
        await state.set_state(CreatorForm.country_other)
        await callback.message.edit_text("3. Укажите страну:")
    await callback.answer()


@router.message(CreatorForm.country_other)
async def creator_country_other(message: Message, state: FSMContext):
    await state.update_data(country=message.text.strip())
    await state.set_state(CreatorForm.region)
    await message.answer("4. Укажите регион / область:")


@router.message(CreatorForm.region)
async def creator_region(message: Message, state: FSMContext):
    await state.update_data(region=message.text.strip())
    await state.set_state(CreatorForm.city)
    await message.answer("Укажите город:")


@router.message(CreatorForm.city)
async def creator_city(message: Message, state: FSMContext):
    await state.update_data(city=message.text.strip())
    await state.set_state(CreatorForm.work_format)
    await message.answer(
        "5. <b>В каком формате вы готовы сотрудничать?</b>",
        reply_markup=single_choice_kb("blog_work", ["Онлайн", "Оффлайн", "Онлайн + оффлайн"])
    )


@router.callback_query(CreatorForm.work_format, F.data.startswith("blog_work:"))
async def creator_work_format(callback: CallbackQuery, state: FSMContext):
    options = ["Онлайн", "Оффлайн", "Онлайн + оффлайн"]
    choice = options[int(callback.data.split(":")[1])]
    await state.update_data(work_format=choice)
    if choice in {"Оффлайн", "Онлайн + оффлайн"}:
        await state.set_state(CreatorForm.travel_scope)
        await callback.message.edit_text(
            "Куда вы готовы выезжать?",
            reply_markup=single_choice_kb("blog_travel", ["Только мой город", "По региону", "По России"])
        )
    else:
        await state.update_data(travel_scope="Не требуется")
        await state.set_state(CreatorForm.niche)
        await state.update_data(niche_selected=[])
        await callback.message.edit_text(
            "6. <b>Тематика вашего блога</b>\nМожно выбрать несколько вариантов.",
            reply_markup=multi_choice_kb("blog_niche", BLOG_NICHES, [])
        )
    await callback.answer()


@router.callback_query(CreatorForm.travel_scope, F.data.startswith("blog_travel:"))
async def creator_travel(callback: CallbackQuery, state: FSMContext):
    options = ["Только мой город", "По региону", "По России"]
    choice = options[int(callback.data.split(":")[1])]
    await state.update_data(travel_scope=choice, niche_selected=[])
    await state.set_state(CreatorForm.niche)
    await callback.message.edit_text(
        "6. <b>Тематика вашего блога</b>\nМожно выбрать несколько вариантов.",
        reply_markup=multi_choice_kb("blog_niche", BLOG_NICHES, [])
    )
    await callback.answer()


@router.callback_query(CreatorForm.niche, F.data.startswith("blog_niche:"))
async def creator_niche(callback: CallbackQuery, state: FSMContext):
    value = callback.data.split(":")[1]
    data = await state.get_data()
    selected = list(data.get("niche_selected", []))
    if value == "done":
        if not selected:
            await callback.answer("Выберите хотя бы один вариант", show_alert=True)
            return
        await state.update_data(niche=", ".join(selected))
        await state.set_state(CreatorForm.blog1_platform)
        await callback.message.edit_text(
            "7. <b>Блог №1</b>\n\nВыберите площадку:",
            reply_markup=single_choice_kb("blog1_platform", BLOG_PLATFORMS)
        )
        await callback.answer()
        return
    option = BLOG_NICHES[int(value)]
    selected.remove(option) if option in selected else selected.append(option)
    await state.update_data(niche_selected=selected)
    await callback.message.edit_reply_markup(reply_markup=multi_choice_kb("blog_niche", BLOG_NICHES, selected))
    await callback.answer()


@router.callback_query(CreatorForm.blog1_platform, F.data.startswith("blog1_platform:"))
async def creator_blog1_platform(callback: CallbackQuery, state: FSMContext):
    platform = BLOG_PLATFORMS[int(callback.data.split(":")[1])]
    await state.update_data(blog1_platform=platform)
    await state.set_state(CreatorForm.blog1_link)
    await callback.message.edit_text("Пришлите ссылку на блог №1:")
    await callback.answer()


@router.message(CreatorForm.blog1_link)
async def creator_blog1_link(message: Message, state: FSMContext):
    await state.update_data(social_link=message.text.strip())
    await state.set_state(CreatorForm.blog1_followers)
    await message.answer("Количество подписчиков в блоге №1:")


@router.message(CreatorForm.blog1_followers)
async def creator_blog1_followers(message: Message, state: FSMContext):
    await state.update_data(followers=message.text.strip())
    await state.set_state(CreatorForm.blog1_reach)
    await message.answer("Средний охват / просмотры блога №1:")


@router.message(CreatorForm.blog1_reach)
async def creator_blog1_reach(message: Message, state: FSMContext):
    await state.update_data(reach=message.text.strip())
    await state.set_state(CreatorForm.add_blog2)
    await message.answer(
        "<b>Добавить второй блог?</b>",
        reply_markup=single_choice_kb("blog2_add", ["Да", "Нет"])
    )


@router.callback_query(CreatorForm.add_blog2, F.data.startswith("blog2_add:"))
async def creator_add_blog2(callback: CallbackQuery, state: FSMContext):
    choice = ["Да", "Нет"][int(callback.data.split(":")[1])]
    if choice == "Да":
        await state.set_state(CreatorForm.blog2_platform)
        await callback.message.edit_text(
            "<b>Блог №2</b>\n\nВыберите площадку:",
            reply_markup=single_choice_kb("blog2_platform", BLOG_PLATFORMS)
        )
    else:
        await state.update_data(
            blog2_platform=None, blog2_link=None, blog2_followers=None, blog2_reach=None,
            ad_formats_selected=[]
        )
        await state.set_state(CreatorForm.ad_formats)
        await callback.message.edit_text(
            "8. <b>Какие рекламные форматы вы размещаете?</b>\nМожно выбрать несколько вариантов.",
            reply_markup=multi_choice_kb("blog_ads", AD_FORMATS, [])
        )
    await callback.answer()


@router.callback_query(CreatorForm.blog2_platform, F.data.startswith("blog2_platform:"))
async def creator_blog2_platform(callback: CallbackQuery, state: FSMContext):
    platform = BLOG_PLATFORMS[int(callback.data.split(":")[1])]
    await state.update_data(blog2_platform=platform)
    await state.set_state(CreatorForm.blog2_link)
    await callback.message.edit_text("Пришлите ссылку на блог №2:")
    await callback.answer()


@router.message(CreatorForm.blog2_link)
async def creator_blog2_link(message: Message, state: FSMContext):
    await state.update_data(blog2_link=message.text.strip())
    await state.set_state(CreatorForm.blog2_followers)
    await message.answer("Количество подписчиков в блоге №2:")


@router.message(CreatorForm.blog2_followers)
async def creator_blog2_followers(message: Message, state: FSMContext):
    await state.update_data(blog2_followers=message.text.strip())
    await state.set_state(CreatorForm.blog2_reach)
    await message.answer("Средний охват / просмотры блога №2:")


@router.message(CreatorForm.blog2_reach)
async def creator_blog2_reach(message: Message, state: FSMContext):
    await state.update_data(blog2_reach=message.text.strip(), ad_formats_selected=[])
    await state.set_state(CreatorForm.ad_formats)
    await message.answer(
        "8. <b>Какие рекламные форматы вы размещаете?</b>\nМожно выбрать несколько вариантов.",
        reply_markup=multi_choice_kb("blog_ads", AD_FORMATS, [])
    )


@router.callback_query(CreatorForm.ad_formats, F.data.startswith("blog_ads:"))
async def creator_ad_formats(callback: CallbackQuery, state: FSMContext):
    value = callback.data.split(":")[1]
    data = await state.get_data()
    selected = list(data.get("ad_formats_selected", []))
    if value == "done":
        if not selected:
            await callback.answer("Выберите хотя бы один вариант", show_alert=True)
            return
        await state.update_data(ad_formats=", ".join(selected), coop_selected=[])
        await state.set_state(CreatorForm.cooperation_formats)
        await callback.message.edit_text(
            "9. <b>Какие варианты сотрудничества рассматриваете?</b>\nМожно выбрать несколько вариантов.",
            reply_markup=multi_choice_kb("blog_coop", COOP_FORMATS, [])
        )
        await callback.answer()
        return
    option = AD_FORMATS[int(value)]
    selected.remove(option) if option in selected else selected.append(option)
    await state.update_data(ad_formats_selected=selected)
    await callback.message.edit_reply_markup(reply_markup=multi_choice_kb("blog_ads", AD_FORMATS, selected))
    await callback.answer()


@router.callback_query(CreatorForm.cooperation_formats, F.data.startswith("blog_coop:"))
async def creator_cooperation(callback: CallbackQuery, state: FSMContext):
    value = callback.data.split(":")[1]
    data = await state.get_data()
    selected = list(data.get("coop_selected", []))
    if value == "done":
        if not selected:
            await callback.answer("Выберите хотя бы один вариант", show_alert=True)
            return
        await state.update_data(cooperation_formats=", ".join(selected))
        await state.set_state(CreatorForm.price_choice)
        await callback.message.edit_text(
            "10. <b>Стоимость сотрудничества</b>",
            reply_markup=single_choice_kb("blog_price", ["Указать стоимость", "Обсуждается"])
        )
        await callback.answer()
        return
    option = COOP_FORMATS[int(value)]
    selected.remove(option) if option in selected else selected.append(option)
    await state.update_data(coop_selected=selected)
    await callback.message.edit_reply_markup(reply_markup=multi_choice_kb("blog_coop", COOP_FORMATS, selected))
    await callback.answer()


@router.callback_query(CreatorForm.price_choice, F.data.startswith("blog_price:"))
async def creator_price_choice(callback: CallbackQuery, state: FSMContext):
    choice = ["Указать стоимость", "Обсуждается"][int(callback.data.split(":")[1])]
    if choice == "Указать стоимость":
        await state.set_state(CreatorForm.price)
        await callback.message.edit_text("Укажите стоимость сотрудничества / стоимость от:")
    else:
        await state.update_data(price="Обсуждается")
        await state.set_state(CreatorForm.excluded_topics)
        await callback.message.edit_text(
            "11. <b>С какими тематиками вы не работаете?</b>\n\n"
            "Напишите ответ или нажмите «Пропустить».",
            reply_markup=single_choice_kb("blog_excluded", ["Пропустить"])
        )
    await callback.answer()


@router.message(CreatorForm.price)
async def creator_price(message: Message, state: FSMContext):
    await state.update_data(price=message.text.strip())
    await state.set_state(CreatorForm.excluded_topics)
    await message.answer(
        "11. <b>С какими тематиками вы не работаете?</b>\n\n"
        "Напишите ответ или нажмите «Пропустить».",
        reply_markup=single_choice_kb("blog_excluded", ["Пропустить"])
    )


@router.callback_query(CreatorForm.excluded_topics, F.data.startswith("blog_excluded:"))
async def creator_excluded_skip(callback: CallbackQuery, state: FSMContext):
    await state.update_data(excluded_topics="Не указано")
    await state.set_state(CreatorForm.brief_ready)
    await callback.message.edit_text(
        "12. <b>Готовы ли вы создавать контент по ТЗ бренда?</b>",
        reply_markup=single_choice_kb("blog_brief", ["Да", "Нет", "Зависит от проекта"])
    )
    await callback.answer()


@router.message(CreatorForm.excluded_topics)
async def creator_excluded(message: Message, state: FSMContext):
    await state.update_data(excluded_topics=message.text.strip())
    await state.set_state(CreatorForm.brief_ready)
    await message.answer(
        "12. <b>Готовы ли вы создавать контент по ТЗ бренда?</b>",
        reply_markup=single_choice_kb("blog_brief", ["Да", "Нет", "Зависит от проекта"])
    )


@router.callback_query(CreatorForm.brief_ready, F.data.startswith("blog_brief:"))
async def creator_brief(callback: CallbackQuery, state: FSMContext):
    options = ["Да", "Нет", "Зависит от проекта"]
    await state.update_data(brief_ready=options[int(callback.data.split(":")[1])])
    await state.set_state(CreatorForm.contact)
    username = f"@{callback.from_user.username}" if callback.from_user.username else "не указан"
    await callback.message.edit_text(
        "13. <b>Контакт для связи</b>\n\n"
        f"Ваш Telegram: {username}\n"
        "Пришлите дополнительный контакт или напишите «Telegram»."
    )
    await callback.answer()


@router.message(CreatorForm.contact)
async def creator_contact(message: Message, state: FSMContext):
    data = await state.get_data()
    upsert_user(message, "creator")
    conn = db()
    conn.execute(
        """
        INSERT INTO creators
        (user_id, name, country, region, city, creator_type, niche, work_format, travel_scope,
         social_link, followers, reach, blog2_platform, blog2_link, blog2_followers, blog2_reach,
         ad_formats, cooperation_formats, price, excluded_topics, brief_ready, contact, status)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'active')
        ON CONFLICT (user_id) DO UPDATE SET
            name=excluded.name, country=excluded.country, region=excluded.region, city=excluded.city,
            creator_type=excluded.creator_type, niche=excluded.niche, work_format=excluded.work_format,
            travel_scope=excluded.travel_scope, social_link=excluded.social_link,
            followers=excluded.followers, reach=excluded.reach,
            blog2_platform=excluded.blog2_platform, blog2_link=excluded.blog2_link,
            blog2_followers=excluded.blog2_followers, blog2_reach=excluded.blog2_reach,
            ad_formats=excluded.ad_formats, cooperation_formats=excluded.cooperation_formats,
            price=excluded.price, excluded_topics=excluded.excluded_topics,
            brief_ready=excluded.brief_ready, contact=excluded.contact, status='active'
        """,
        (
            message.from_user.id, data["name"], data.get("country"), data.get("region"), data.get("city"),
            "Блогер", data.get("niche"), data.get("work_format"), data.get("travel_scope"),
            data.get("social_link"), data.get("followers"), data.get("reach"),
            data.get("blog2_platform"), data.get("blog2_link"), data.get("blog2_followers"), data.get("blog2_reach"),
            data.get("ad_formats"), data.get("cooperation_formats"), data.get("price"),
            data.get("excluded_topics"), data.get("brief_ready"), message.text.strip()
        )
    )
    conn.commit()
    conn.close()
    await state.clear()
    await message.answer(
        "<b>Готово. Профиль блогера создан.</b>\n\n"
        "Теперь можно смотреть актуальные проекты и откликаться на подходящие предложения.",
        reply_markup=creator_menu_kb()
    )



# ---------- Owner businesses ----------

@router.message(F.text == "🏢 Бизнесы")
async def owner_businesses(message: Message):
    if not is_admin(message.from_user.id):
        return
    await message.answer(
        "<b>Бизнесы</b>\n\n"
        "Добавляйте бизнесы вручную и создавайте от их имени заявки.",
        reply_markup=admin_businesses_kb()
    )


@router.callback_query(F.data == "admin_business_add")
async def admin_business_add(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        return
    await state.clear()
    await state.set_state(AdminBusinessForm.business_name)
    await callback.message.edit_text("<b>Добавить бизнес</b>\n\nНазвание бизнеса:")
    await callback.answer()


@router.callback_query(F.data.startswith("admin_business_edit:"))
async def admin_business_edit(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        return
    business_id = int(callback.data.split(":")[1])
    await state.clear()
    await state.update_data(edit_business_id=business_id)
    await state.set_state(AdminBusinessForm.business_name)
    await callback.message.answer(
        "<b>Редактировать бизнес</b>\n\n"
        "Введите данные заново — они заменят текущие.\n\nНазвание бизнеса:"
    )
    await callback.answer()


@router.message(AdminBusinessForm.business_name)
async def admin_business_name(message: Message, state: FSMContext):
    await state.update_data(business_name=message.text.strip())
    await state.set_state(AdminBusinessForm.city)
    await message.answer("Город / онлайн:")


@router.message(AdminBusinessForm.city)
async def admin_business_city(message: Message, state: FSMContext):
    await state.update_data(city=message.text.strip())
    await state.set_state(AdminBusinessForm.niche)
    await message.answer("Ниша бизнеса:")


@router.message(AdminBusinessForm.niche)
async def admin_business_niche(message: Message, state: FSMContext):
    await state.update_data(niche=message.text.strip())
    await state.set_state(AdminBusinessForm.social_link)
    await message.answer("Ссылка на соцсети бизнеса:")


@router.message(AdminBusinessForm.social_link)
async def admin_business_social(message: Message, state: FSMContext):
    await state.update_data(social_link=message.text.strip())
    await state.set_state(AdminBusinessForm.contact)
    await message.answer("Контакт для связи:")


@router.message(AdminBusinessForm.contact)
async def admin_business_contact(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    data = await state.get_data()
    conn = db()
    edit_id = data.get("edit_business_id")
    if edit_id:
        conn.execute(
            """
            UPDATE admin_businesses
            SET business_name=?, city=?, niche=?, social_link=?, contact=?
            WHERE business_id=?
            """,
            (
                data["business_name"], data["city"], data["niche"],
                data["social_link"], message.text.strip(), edit_id
            )
        )
        result_text = "<b>Карточка бизнеса обновлена.</b>"
    else:
        conn.execute(
            """
            INSERT INTO admin_businesses
            (business_name, city, niche, social_link, contact, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                data["business_name"], data["city"], data["niche"],
                data["social_link"], message.text.strip(), datetime.now().isoformat()
            )
        )
        result_text = "<b>Бизнес добавлен.</b>"
    conn.commit()
    conn.close()
    await state.clear()
    await message.answer(result_text, reply_markup=admin_businesses_kb())


@router.callback_query(F.data == "admin_business_list")
async def admin_business_list(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        return
    conn = db()
    rows = conn.execute(
        "SELECT * FROM admin_businesses ORDER BY business_id DESC LIMIT 30"
    ).fetchall()
    conn.close()

    if not rows:
        await callback.message.edit_text(
            "Бизнесов пока нет.",
            reply_markup=admin_businesses_kb()
        )
        await callback.answer()
        return

    await callback.message.edit_text("<b>Все бизнесы</b>")
    for b in rows:
        await callback.message.answer(
            f"<b>{b['business_name']}</b>\n"
            f"Город: {b['city'] or '—'}\n"
            f"Ниша: {b['niche'] or '—'}\n"
            f"Соцсети: {b['social_link'] or '—'}\n"
            f"Контакт: {b['contact'] or '—'}",
            reply_markup=admin_business_card_kb(b["business_id"])
        )
    await callback.answer()


@router.callback_query(F.data == "admin_request_choose_business")
async def admin_request_choose_business(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        return
    conn = db()
    rows = conn.execute(
        "SELECT business_id, business_name FROM admin_businesses ORDER BY business_name"
    ).fetchall()
    conn.close()

    if not rows:
        await callback.message.edit_text(
            "Сначала добавьте хотя бы один бизнес.",
            reply_markup=admin_businesses_kb()
        )
        await callback.answer()
        return

    kb = InlineKeyboardBuilder()
    for b in rows:
        kb.button(
            text=b["business_name"],
            callback_data=f"admin_request_for:{b['business_id']}"
        )
    kb.adjust(1)
    await callback.message.edit_text("<b>Выберите бизнес</b>", reply_markup=kb.as_markup())
    await callback.answer()


@router.callback_query(F.data.startswith("admin_request_for:"))
async def admin_request_for_business(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        return
    business_id = int(callback.data.split(":")[1])

    conn = db()
    business = conn.execute(
        "SELECT * FROM admin_businesses WHERE business_id=?",
        (business_id,)
    ).fetchone()
    conn.close()

    if not business:
        await callback.answer("Бизнес не найден", show_alert=True)
        return

    await state.clear()
    await state.update_data(admin_business_id=business_id)
    await state.set_state(AdminRequestForm.title)
    await callback.message.answer(
        f"<b>Новая заявка от {business['business_name']}</b>\n\n"
        "Короткий заголовок заявки:"
    )
    await callback.answer()


@router.message(AdminRequestForm.title)
async def admin_request_title(message: Message, state: FSMContext):
    await state.update_data(title=message.text.strip())
    await state.set_state(AdminRequestForm.city)
    await message.answer("Город / география проекта:")


@router.message(AdminRequestForm.city)
async def admin_request_city(message: Message, state: FSMContext):
    await state.update_data(city=message.text.strip())
    await state.set_state(AdminRequestForm.task)
    await message.answer("Что нужно сделать блогеру?")


@router.message(AdminRequestForm.task)
async def admin_request_task(message: Message, state: FSMContext):
    await state.update_data(task=message.text.strip())
    await state.set_state(AdminRequestForm.cooperation_format)
    await message.answer("Формат сотрудничества: оплата / бартер / % / другое:")


@router.message(AdminRequestForm.cooperation_format)
async def admin_request_cooperation(message: Message, state: FSMContext):
    await state.update_data(cooperation_format=message.text.strip())
    await state.set_state(AdminRequestForm.budget)
    await message.answer("Бюджет / условия:")


@router.message(AdminRequestForm.budget)
async def admin_request_budget(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return

    data = await state.get_data()
    conn = db()
    business = conn.execute(
        "SELECT * FROM admin_businesses WHERE business_id=?",
        (data["admin_business_id"],)
    ).fetchone()

    if not business:
        conn.close()
        await state.clear()
        await message.answer("Бизнес не найден.", reply_markup=owner_reply_kb())
        return

    full_task = f"{data['title']}\n\n{data['task']}"
    row = conn.execute(
        """
        INSERT INTO requests
        (business_user_id, business_name, city, niche, task, creator_needed,
         cooperation_format, budget, social_link, contact, status, created_at,
         moderated_at, moderated_by)
        VALUES (?, ?, ?, ?, ?, 'Блогер', ?, ?, ?, ?, 'active', ?, ?, ?)
        RETURNING request_id
        """,
        (
            message.from_user.id,
            business["business_name"],
            data["city"],
            business["niche"],
            full_task,
            data["cooperation_format"],
            message.text.strip(),
            business["social_link"],
            business["contact"],
            datetime.now().isoformat(),
            datetime.now().isoformat(),
            message.from_user.id
        )
    ).fetchone()
    request_id = row["request_id"]
    conn.commit()
    conn.close()
    await state.clear()

    await message.answer(
        f"<b>Заявка опубликована.</b>\n\n"
        f"Бизнес: {business['business_name']}\n"
        f"Заявка #{request_id}\n\n"
        "Она уже доступна блогерам в разделе «Проекты».",
        reply_markup=owner_reply_kb()
    )

# ---------- Profile editing ----------

@router.callback_query(F.data == "creator_edit_profile")
async def creator_edit_profile(callback: CallbackQuery, state: FSMContext):
    conn = db()
    creator = conn.execute(
        "SELECT creator_type FROM creators WHERE user_id=?",
        (callback.from_user.id,)
    ).fetchone()
    conn.close()

    await state.clear()
    if creator and creator["creator_type"] == "Контент-креатор":
        await state.set_state(ContentCreatorForm.name)
        await callback.message.edit_text(
            "<b>Редактирование профиля контент-креатора</b>\n\n"
            "Пройдите анкету заново — новые ответы заменят старые.\n\n"
            "1. Как вас зовут?"
        )
    else:
        await state.update_data(creator_type="Блогер")
        await state.set_state(CreatorForm.name)
        await callback.message.edit_text(
            "<b>Редактирование профиля блогера</b>\n\n"
            "Пройдите анкету заново — новые ответы заменят старые.\n\n"
            "1. Как вас зовут?"
        )
    await callback.answer()


@router.callback_query(F.data == "business_edit_profile")
async def business_edit_profile(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    await state.set_state(BusinessForm.business_name)
    await callback.message.edit_text(
        "<b>Редактирование профиля бизнеса</b>\n\n"
        "Пройдите анкету заново — новые ответы заменят старые.\n\n"
        "Введите название бизнеса:"
    )
    await callback.answer()


# ---------- Business request creation ----------

@router.callback_query(F.data == "business_create_request")
async def create_request_start(callback: CallbackQuery, state: FSMContext):
    conn = db()
    profile = conn.execute("SELECT * FROM business_profiles WHERE user_id=?", (callback.from_user.id,)).fetchone()
    conn.close()
    if not profile:
        await callback.message.edit_text("Сначала заполните профиль бизнеса.", reply_markup=main_menu_kb())
        await callback.answer()
        return
    await state.set_state(RequestForm.creator_needed)
    await callback.message.edit_text("Кого ищем? Например: UGC-креатор, блогер, фотограф, видеограф")
    await callback.answer()


@router.message(RequestForm.creator_needed)
async def req_creator_needed(message: Message, state: FSMContext):
    await state.update_data(creator_needed=message.text)
    await state.set_state(RequestForm.task)
    await message.answer("Что нужно сделать? Опишите задачу:")


@router.message(RequestForm.task)
async def req_task(message: Message, state: FSMContext):
    await state.update_data(task=message.text)
    await state.set_state(RequestForm.cooperation_format)
    await message.answer("Формат сотрудничества: оплата / бартер / процент / коллаборация / обсудим")


@router.message(RequestForm.cooperation_format)
async def req_format(message: Message, state: FSMContext):
    await state.update_data(cooperation_format=message.text)
    await state.set_state(RequestForm.budget)
    await message.answer("Бюджет или что даёте по бартеру:")


@router.message(RequestForm.budget)
async def req_budget(message: Message, state: FSMContext, bot: Bot):
    data = await state.get_data()
    conn = db()
    profile = conn.execute("SELECT * FROM business_profiles WHERE user_id=?", (message.from_user.id,)).fetchone()
    row = conn.execute(
        """
        INSERT INTO requests
        (business_user_id, business_name, city, niche, task, creator_needed, cooperation_format, budget, social_link, contact, status, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'pending', ?)
        RETURNING request_id
        """,
        (
            message.from_user.id, profile["business_name"], profile["city"], profile["niche"],
            data["task"], data["creator_needed"], data["cooperation_format"], message.text,
            profile["social_link"], profile["contact"], datetime.now().isoformat()
        )
    ).fetchone()
    request_id = row["request_id"]
    request = conn.execute("SELECT * FROM requests WHERE request_id=?", (request_id,)).fetchone()
    conn.commit()
    conn.close()

    await state.clear()
    await message.answer(
        "<b>Заявка отправлена на модерацию.</b>\n\n"
        "После одобрения она появится у креаторов.",
        reply_markup=business_menu_kb(),
    )

    for admin_id in ADMIN_IDS:
        try:
            await bot.send_message(
                admin_id,
                "<b>Новая заявка на модерацию</b>\n\n" + request_card(request),
                reply_markup=moderation_kb(request_id),
            )
        except Exception:
            pass


# ---------- Creator views/responds ----------

@router.callback_query(F.data == "creator_view_requests")
async def creator_view_requests(callback: CallbackQuery):
    conn = db()
    requests = conn.execute("SELECT * FROM requests WHERE status='active' ORDER BY request_id DESC LIMIT 10").fetchall()
    conn.close()
    if not requests:
        await callback.message.edit_text("Сейчас активных заявок нет.", reply_markup=creator_menu_kb())
        await callback.answer()
        return
    await callback.message.edit_text("Активные заявки:")
    for r in requests:
        await callback.message.answer(request_card(r), reply_markup=request_kb(r["request_id"]))
    await callback.answer()


@router.callback_query(F.data.startswith("respond:"))
async def respond_to_request(callback: CallbackQuery, bot: Bot):
    request_id = int(callback.data.split(":")[1])
    conn = db()
    creator = conn.execute("SELECT * FROM creators WHERE user_id=?", (callback.from_user.id,)).fetchone()
    if not creator:
        conn.close()
        await callback.answer("Сначала заполните анкету креатора", show_alert=True)
        return

    request = conn.execute("SELECT * FROM requests WHERE request_id=? AND status='active'", (request_id,)).fetchone()
    if not request:
        conn.close()
        await callback.answer("Заявка уже не активна", show_alert=True)
        return

    try:
        row = conn.execute(
            "INSERT INTO responses(request_id, creator_user_id, business_user_id, status, created_at) VALUES (?, ?, ?, 'new', ?) RETURNING response_id",
            (request_id, callback.from_user.id, request["business_user_id"], datetime.now().isoformat())
        ).fetchone()
        response_id = row["response_id"]
        conn.commit()
    except (sqlite3.IntegrityError, psycopg.IntegrityError):
        conn.close()
        await callback.answer("Вы уже откликались на эту заявку", show_alert=True)
        return
    conn.close()

    await callback.answer("Отклик отправлен")
    await callback.message.answer("Ваш отклик отправлен бизнесу.")

    await bot.send_message(
        request["business_user_id"],
        "<b>Новый отклик на вашу заявку</b>\n\n"
        + request_card(request)
        + "\n\n"
        + creator_card(creator),
        reply_markup=business_decision_kb(response_id)
    )


# ---------- Business accepts/declines ----------

@router.callback_query(F.data.startswith("accept:"))
async def accept_response(callback: CallbackQuery, bot: Bot):
    response_id = int(callback.data.split(":")[1])
    conn = db()
    response = conn.execute("SELECT * FROM responses WHERE response_id=?", (response_id,)).fetchone()
    if not response or response["business_user_id"] != callback.from_user.id:
        conn.close()
        await callback.answer("Отклик не найден", show_alert=True)
        return

    conn.execute("UPDATE responses SET status='accepted' WHERE response_id=?", (response_id,))
    creator = conn.execute("SELECT * FROM creators WHERE user_id=?", (response["creator_user_id"],)).fetchone()
    request = conn.execute("SELECT * FROM requests WHERE request_id=?", (response["request_id"],)).fetchone()
    business = conn.execute("SELECT * FROM business_profiles WHERE user_id=?", (response["business_user_id"],)).fetchone()
    conn.commit()
    conn.close()

    await callback.message.edit_reply_markup(reply_markup=None)
    await callback.message.answer(
        "Отклик принят. Контакты блогера:\n\n" + contact_line(creator["user_id"], creator["contact"])
    )

    if business:
        business_contacts = contact_line(business["user_id"], business["contact"])
    else:
        # Для бизнеса, добавленного владельцем вручную.
        business_contacts = request["contact"] or "Представитель проекта свяжется с вами."

    await bot.send_message(
        creator["user_id"],
        "Ваш отклик принят 🎉\n\n"
        f"Бизнес: {request['business_name']}\n"
        f"Заявка: {request['task']}\n\n"
        "Контакты бизнеса:\n"
        + business_contacts
    )
    await callback.answer("Принято")


@router.callback_query(F.data.startswith("decline:"))
async def decline_response(callback: CallbackQuery, bot: Bot):
    response_id = int(callback.data.split(":")[1])
    conn = db()
    response = conn.execute("SELECT * FROM responses WHERE response_id=?", (response_id,)).fetchone()
    if not response or response["business_user_id"] != callback.from_user.id:
        conn.close()
        await callback.answer("Отклик не найден", show_alert=True)
        return

    conn.execute("UPDATE responses SET status='declined' WHERE response_id=?", (response_id,))
    request = conn.execute("SELECT * FROM requests WHERE request_id=?", (response["request_id"],)).fetchone()
    conn.commit()
    conn.close()

    await callback.message.edit_reply_markup(reply_markup=None)
    await callback.message.answer("Отклик отклонён.")
    await bot.send_message(
        response["creator_user_id"],
        f"По заявке #{request['request_id']} бизнес выбрал другого креатора. Новые предложения появятся в боте."
    )
    await callback.answer("Отклонено")


# ---------- Lists ----------

@router.callback_query(F.data == "business_my_requests")
async def business_my_requests(callback: CallbackQuery):
    conn = db()
    rows = conn.execute(
        "SELECT * FROM requests WHERE business_user_id=? ORDER BY request_id DESC LIMIT 10",
        (callback.from_user.id,)
    ).fetchall()
    conn.close()
    if not rows:
        await callback.message.edit_text("У вас пока нет заявок.", reply_markup=business_menu_kb())
        await callback.answer()
        return
    await callback.message.edit_text("Ваши заявки:")
    for r in rows:
        await callback.message.answer(request_card(r))
    await callback.answer()


@router.callback_query(F.data == "business_responses")
async def business_responses(callback: CallbackQuery):
    conn = db()
    rows = conn.execute(
        """
        SELECT responses.*, creators.name, creators.social_link, requests.task
        FROM responses
        JOIN creators ON creators.user_id = responses.creator_user_id
        JOIN requests ON requests.request_id = responses.request_id
        WHERE responses.business_user_id=?
        ORDER BY responses.response_id DESC LIMIT 10
        """,
        (callback.from_user.id,)
    ).fetchall()
    conn.close()
    if not rows:
        await callback.message.edit_text("Откликов пока нет.", reply_markup=business_menu_kb())
        await callback.answer()
        return
    await callback.message.edit_text("Последние отклики:")
    for r in rows:
        await callback.message.answer(
            f"Заявка: {r['task']}\nКреатор: {r['name']}\nСоцсети: {r['social_link']}\nСтатус: {r['status']}",
            reply_markup=business_decision_kb(r["response_id"]) if r["status"] == "new" else None
        )
    await callback.answer()


@router.callback_query(F.data == "creator_my_responses")
async def creator_my_responses(callback: CallbackQuery):
    conn = db()
    rows = conn.execute(
        """
        SELECT responses.*, requests.business_name, requests.task
        FROM responses
        JOIN requests ON requests.request_id = responses.request_id
        WHERE responses.creator_user_id=?
        ORDER BY responses.response_id DESC LIMIT 10
        """,
        (callback.from_user.id,)
    ).fetchall()
    conn.close()
    if not rows:
        await callback.message.edit_text("У вас пока нет откликов.", reply_markup=creator_menu_kb())
        await callback.answer()
        return
    await callback.message.edit_text("Ваши отклики:")
    for r in rows:
        await callback.message.answer(
            f"Бизнес: {r['business_name']}\nЗадача: {r['task']}\nСтатус: {r['status']}"
        )
    await callback.answer()


# ---------- Referrals ----------

@router.callback_query(F.data == "my_referral")
async def my_referral(callback: CallbackQuery, bot: Bot):
    me = await bot.get_me()
    counts = referral_counts(callback.from_user.id)
    link = f"https://t.me/{me.username}?start=ref_{callback.from_user.id}"

    conn = db()
    user = conn.execute("SELECT role FROM users WHERE user_id=?", (callback.from_user.id,)).fetchone()
    conn.close()
    role = user["role"] if user else None
    menu = business_menu_kb() if role == "business" else creator_menu_kb()

    await callback.message.edit_text(
        "<b>Пригласить в КЛИК</b>\n\n"
        "Отправьте свою персональную ссылку друзьям, коллегам, блогерам и бизнесам. "
        "Мы сохраним, кого вы пригласили.\n\n"
        f"<b>Ваша ссылка:</b>\n<code>{link}</code>\n\n"
        f"Приглашено: <b>{counts['total']}</b>\n"
        f"Креаторов: <b>{counts['creators']}</b>\n"
        f"Бизнесов: <b>{counts['businesses']}</b>",
        reply_markup=menu,
    )
    await callback.answer()


# ---------- Profiles / help ----------

@router.callback_query(F.data == "business_profile")
async def business_profile(callback: CallbackQuery):
    conn = db()
    p = conn.execute("SELECT * FROM business_profiles WHERE user_id=?", (callback.from_user.id,)).fetchone()
    conn.close()
    if not p:
        await callback.message.edit_text("Профиль бизнеса пока не заполнен.", reply_markup=business_intro_kb())
    else:
        await callback.message.edit_text(
            "<b>Мой профиль</b>\n\n"
            f"Бизнес: {p['business_name']}\n"
            f"Город: {p['city']}\n"
            f"Ниша: {p['niche']}\n"
            f"Соцсети: {p['social_link']}\n"
            f"Контакт: {p['contact']}",
            reply_markup=business_menu_kb()
        )
    await callback.answer()


@router.callback_query(F.data == "creator_profile")
async def creator_profile(callback: CallbackQuery):
    conn = db()
    c = conn.execute("SELECT * FROM creators WHERE user_id=?", (callback.from_user.id,)).fetchone()
    conn.close()
    if not c:
        await callback.message.edit_text("Профиль креатора пока не заполнен.", reply_markup=creator_intro_kb())
    else:
        await callback.message.edit_text(creator_card(c), reply_markup=creator_menu_kb())
    await callback.answer()


@router.callback_query(F.data == "creator_settings")
async def creator_settings(callback: CallbackQuery):
    await callback.message.edit_text(
        "<b>Настройки</b>\n\nСкоро здесь появятся настройки города, уведомлений и форматов сотрудничества.",
        reply_markup=creator_menu_kb()
    )
    await callback.answer()


@router.callback_query(F.data.in_({"help_creator", "help_business"}))
async def help_section(callback: CallbackQuery):
    is_business = callback.data == "help_business"
    await callback.message.edit_text(
        "<b>Помощь</b>\n\n"
        "КЛИК соединяет бизнес и креаторов.\n\n"
        "Если что-то не получается, напишите администратору проекта. Контакт добавим перед публичным запуском.",
        reply_markup=business_menu_kb() if is_business else creator_menu_kb()
    )
    await callback.answer()

# ---------- Owner / Admin ----------

async def send_pending_requests(message: Message):
    conn = db()
    rows = conn.execute("SELECT * FROM requests WHERE status='pending' ORDER BY request_id ASC LIMIT 20").fetchall()
    conn.close()
    if not rows:
        await message.answer("На модерации сейчас ничего нет.", reply_markup=owner_reply_kb())
        return
    await message.answer(f"<b>На модерации: {len(rows)}</b>", reply_markup=owner_reply_kb())
    for r in rows:
        await message.answer(request_card(r), reply_markup=moderation_kb(r["request_id"]))


async def broadcast_approved_request(bot: Bot, request_id: int):
    conn = db()
    request = conn.execute("SELECT * FROM requests WHERE request_id=?", (request_id,)).fetchone()
    creators = conn.execute("SELECT * FROM creators WHERE status='active'").fetchall()
    conn.close()
    if not request or request["status"] != "active":
        return
    for creator in creators:
        try:
            await bot.send_message(
                creator["user_id"],
                "<b>Новая заявка</b>\n\n" + request_card(request),
                reply_markup=request_kb(request_id),
            )
        except Exception:
            pass


@router.message(Command("admin"))
async def admin(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("Нет доступа.")
        return
    await message.answer("<b>Панель владельца КЛИК</b>", reply_markup=owner_reply_kb())


@router.message(F.text == "🛡 Модерация")
async def owner_moderation_message(message: Message):
    if not is_admin(message.from_user.id):
        return
    await send_pending_requests(message)


@router.callback_query(F.data == "admin_moderation")
async def owner_moderation_callback(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await send_pending_requests(callback.message)
    await callback.answer()


@router.callback_query(F.data.startswith("moderate_approve:"))
async def moderate_approve(callback: CallbackQuery, bot: Bot):
    if not is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    request_id = int(callback.data.split(":")[1])
    conn = db()
    request = conn.execute("SELECT * FROM requests WHERE request_id=?", (request_id,)).fetchone()
    if not request or request["status"] != "pending":
        conn.close()
        await callback.answer("Заявка уже обработана", show_alert=True)
        return
    conn.execute(
        "UPDATE requests SET status='active', moderated_at=?, moderated_by=? WHERE request_id=?",
        (datetime.now().isoformat(), callback.from_user.id, request_id),
    )
    conn.commit()
    conn.close()
    await callback.message.edit_reply_markup(reply_markup=None)
    await callback.message.answer(f"✅ Заявка #{request_id} одобрена и опубликована.")
    try:
        await bot.send_message(
            request["business_user_id"],
            f"✅ Ваша заявка #{request_id} прошла модерацию и опубликована.",
            reply_markup=business_menu_kb(),
        )
    except Exception:
        pass
    await broadcast_approved_request(bot, request_id)
    await callback.answer("Опубликовано")


@router.callback_query(F.data.startswith("moderate_reject:"))
async def moderate_reject(callback: CallbackQuery, bot: Bot):
    if not is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    request_id = int(callback.data.split(":")[1])
    conn = db()
    request = conn.execute("SELECT * FROM requests WHERE request_id=?", (request_id,)).fetchone()
    if not request or request["status"] != "pending":
        conn.close()
        await callback.answer("Заявка уже обработана", show_alert=True)
        return
    conn.execute(
        "UPDATE requests SET status='rejected', moderated_at=?, moderated_by=? WHERE request_id=?",
        (datetime.now().isoformat(), callback.from_user.id, request_id),
    )
    conn.commit()
    conn.close()
    await callback.message.edit_reply_markup(reply_markup=None)
    await callback.message.answer(f"❌ Заявка #{request_id} отклонена.")
    try:
        await bot.send_message(
            request["business_user_id"],
            f"Заявка #{request_id} не прошла модерацию. Проверьте формулировку и создайте новую заявку.",
            reply_markup=business_menu_kb(),
        )
    except Exception:
        pass
    await callback.answer("Отклонено")


async def stats_text() -> str:
    conn = db()
    users = conn.execute("SELECT COUNT(*) AS c FROM users").fetchone()["c"]
    creators = conn.execute("SELECT COUNT(*) AS c FROM creators").fetchone()["c"]
    businesses = conn.execute("SELECT COUNT(*) AS c FROM business_profiles").fetchone()["c"]
    requests = conn.execute("SELECT COUNT(*) AS c FROM requests").fetchone()["c"]
    pending = conn.execute("SELECT COUNT(*) AS c FROM requests WHERE status='pending'").fetchone()["c"]
    active = conn.execute("SELECT COUNT(*) AS c FROM requests WHERE status='active'").fetchone()["c"]
    responses = conn.execute("SELECT COUNT(*) AS c FROM responses").fetchone()["c"]
    accepted = conn.execute("SELECT COUNT(*) AS c FROM responses WHERE status='accepted'").fetchone()["c"]
    referred = conn.execute("SELECT COUNT(*) AS c FROM users WHERE referred_by IS NOT NULL").fetchone()["c"]
    referrers = conn.execute("SELECT COUNT(DISTINCT referred_by) AS c FROM users WHERE referred_by IS NOT NULL").fetchone()["c"]
    conn.close()
    return (
        "<b>Статистика КЛИК</b>\n\n"
        f"Пользователи: {users}\n"
        f"Креаторы: {creators}\n"
        f"Бизнесы: {businesses}\n\n"
        f"Заявки: {requests}\n"
        f"На модерации: {pending}\n"
        f"Активные: {active}\n\n"
        f"Отклики: {responses}\n"
        f"Приняты: {accepted}\n\n"
        f"Пришли по рефералам: {referred}\n"
        f"Приглашают пользователей: {referrers}"
    )


@router.message(F.text == "📊 Статистика")
async def owner_stats_message(message: Message):
    if is_admin(message.from_user.id):
        await message.answer(await stats_text(), reply_markup=owner_reply_kb())


@router.callback_query(F.data == "admin_stats")
async def admin_stats(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await callback.message.answer(await stats_text(), reply_markup=owner_reply_kb())
    await callback.answer()


@router.message(F.text.in_({"👥 База креаторов", "👥 Медиа-база"}))
async def owner_creators(message: Message):
    if not is_admin(message.from_user.id):
        return
    conn = db()
    rows = conn.execute("SELECT * FROM creators ORDER BY user_id DESC LIMIT 30").fetchall()
    conn.close()
    if not rows:
        await message.answer("Креаторов пока нет.", reply_markup=owner_reply_kb())
        return
    await message.answer(f"<b>Медиа-база — {len(rows)} последних</b>", reply_markup=owner_reply_kb())
    for c in rows:
        await message.answer(creator_card(c) + "\n\n" + contact_line(c["user_id"], c["contact"]))


@router.message(F.text == "🏢 База бизнесов")
async def owner_businesses(message: Message):
    if not is_admin(message.from_user.id):
        return
    conn = db()
    rows = conn.execute("SELECT * FROM business_profiles ORDER BY user_id DESC LIMIT 30").fetchall()
    conn.close()
    if not rows:
        await message.answer("Бизнесов пока нет.", reply_markup=owner_reply_kb())
        return
    await message.answer(f"<b>Бизнесы — {len(rows)} последних</b>", reply_markup=owner_reply_kb())
    for p in rows:
        await message.answer(
            f"<b>{p['business_name']}</b>\n"
            f"Город: {p['city']}\nНиша: {p['niche']}\nСоцсети: {p['social_link']}\n"
            + contact_line(p["user_id"], p["contact"])
        )


@router.message(F.text == "📋 Все заявки")
async def owner_requests(message: Message):
    if not is_admin(message.from_user.id):
        return
    conn = db()
    rows = conn.execute("SELECT * FROM requests ORDER BY request_id DESC LIMIT 30").fetchall()
    conn.close()
    if not rows:
        await message.answer("Заявок пока нет.", reply_markup=owner_reply_kb())
        return
    await message.answer(f"<b>Последние заявки: {len(rows)}</b>", reply_markup=owner_reply_kb())
    for r in rows:
        await message.answer(request_card(r), reply_markup=moderation_kb(r["request_id"]) if r["status"] == "pending" else None)


@router.callback_query(F.data == "admin_requests")
async def admin_requests(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await owner_requests(callback.message)
    await callback.answer()


@router.message(F.text == "💬 Все отклики")
async def owner_responses(message: Message):
    if not is_admin(message.from_user.id):
        return
    conn = db()
    rows = conn.execute(
        """
        SELECT responses.*, creators.name, requests.business_name, requests.task
        FROM responses
        JOIN creators ON creators.user_id=responses.creator_user_id
        JOIN requests ON requests.request_id=responses.request_id
        ORDER BY responses.response_id DESC LIMIT 30
        """
    ).fetchall()
    conn.close()
    if not rows:
        await message.answer("Откликов пока нет.", reply_markup=owner_reply_kb())
        return
    await message.answer(f"<b>Последние отклики: {len(rows)}</b>", reply_markup=owner_reply_kb())
    for r in rows:
        markup = business_decision_kb(r["response_id"]) if r["status"] == "new" else None
        await message.answer(
            f"#{r['response_id']} · {r['status']}\n"
            f"Бизнес: {r['business_name']}\n"
            f"Блогер: {r['name']}\n"
            f"Задача: {r['task']}",
            reply_markup=markup
        )


@router.message(F.text == "🔗 Рефералы")
async def owner_referrals(message: Message):
    if not is_admin(message.from_user.id):
        return
    conn = db()
    rows = conn.execute(
        """
        SELECT
            inviter.user_id AS inviter_id,
            inviter.full_name AS inviter_name,
            inviter.username AS inviter_username,
            COUNT(invited.user_id) AS total,
            SUM(CASE WHEN invited.role='creator' THEN 1 ELSE 0 END) AS creators,
            SUM(CASE WHEN invited.role='business' THEN 1 ELSE 0 END) AS businesses
        FROM users inviter
        JOIN users invited ON invited.referred_by = inviter.user_id
        GROUP BY inviter.user_id, inviter.full_name, inviter.username
        ORDER BY total DESC
        LIMIT 30
        """
    ).fetchall()
    conn.close()
    if not rows:
        await message.answer("Рефералов пока нет.", reply_markup=owner_reply_kb())
        return
    await message.answer("<b>Реферальная статистика</b>", reply_markup=owner_reply_kb())
    for r in rows:
        username = f"@{r['inviter_username']}" if r['inviter_username'] else "без username"
        await message.answer(
            f"<b>{r['inviter_name'] or 'Пользователь'}</b> · {username}\n"
            f"Приглашено: {r['total']}\n"
            f"Креаторов: {r['creators'] or 0} · Бизнесов: {r['businesses'] or 0}"
        )


def export_excel_bytes() -> bytes:
    conn = db()
    datasets = {
        "Креаторы": conn.execute(
            "SELECT user_id, name, country, region, city, creator_type, niche, work_format, travel_scope, social_link, followers, reach, blog2_platform, blog2_link, blog2_followers, blog2_reach, ad_formats, content_types, industries, on_camera, creator_skills, portfolio_link, delivery_available, cooperation_formats, price, excluded_topics, brief_ready, contact, status FROM creators ORDER BY user_id"
        ).fetchall(),
        "Бизнес": conn.execute(
            "SELECT user_id, business_name, city, niche, social_link, contact FROM business_profiles ORDER BY user_id"
        ).fetchall(),
        "Заявки": conn.execute(
            "SELECT request_id, business_user_id, business_name, city, niche, creator_needed, task, cooperation_format, budget, social_link, contact, status, created_at FROM requests ORDER BY request_id"
        ).fetchall(),
        "Отклики": conn.execute(
            "SELECT response_id, request_id, creator_user_id, business_user_id, status, created_at FROM responses ORDER BY response_id"
        ).fetchall(),
        "Рефералы": conn.execute(
            """SELECT user_id, username, full_name, role, referred_by, referral_created_at, created_at
            FROM users WHERE referred_by IS NOT NULL ORDER BY referral_created_at"""
        ).fetchall(),
    }
    conn.close()

    wb = Workbook()
    wb.remove(wb.active)
    for title, rows in datasets.items():
        ws = wb.create_sheet(title)
        if not rows:
            ws.append(["Нет данных"])
            continue
        headers = list(rows[0].keys())
        ws.append(headers)
        for row in rows:
            ws.append([row[h] for h in headers])
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="111827")
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"
        for column_cells in ws.columns:
            width = min(max(len(str(c.value or "")) for c in column_cells) + 2, 42)
            ws.column_dimensions[column_cells[0].column_letter].width = max(width, 12)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


@router.message(F.text == "📥 Выгрузить Excel")
async def owner_export_message(message: Message):
    if not is_admin(message.from_user.id):
        return
    data = export_excel_bytes()
    filename = f"KLIK_base_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"
    await message.answer_document(BufferedInputFile(data, filename=filename), caption="Актуальная база КЛИК")


@router.callback_query(F.data == "admin_export")
async def owner_export_callback(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    data = export_excel_bytes()
    filename = f"KLIK_base_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"
    await callback.message.answer_document(BufferedInputFile(data, filename=filename), caption="Актуальная база КЛИК")
    await callback.answer()


@router.message(F.text == "📣 Рассылка")
async def owner_broadcast_start(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    await state.set_state(BroadcastForm.text)
    await message.answer(
        "Пришлите текст рассылки одним сообщением. Он уйдёт всем зарегистрированным пользователям.\n\n"
        "Для отмены отправьте /cancel."
    )


@router.message(Command("cancel"))
async def cancel_action(message: Message, state: FSMContext):
    await state.clear()
    if is_admin(message.from_user.id):
        await message.answer("Действие отменено.", reply_markup=owner_reply_kb())
    else:
        await message.answer("Действие отменено.")


@router.message(BroadcastForm.text)
async def owner_broadcast_send(message: Message, state: FSMContext, bot: Bot):
    if not is_admin(message.from_user.id):
        await state.clear()
        return
    text = message.html_text or message.text or ""
    conn = db()
    users = conn.execute("SELECT user_id FROM users ORDER BY user_id").fetchall()
    conn.close()
    sent = 0
    failed = 0
    for row in users:
        try:
            await bot.send_message(row["user_id"], text)
            sent += 1
        except Exception:
            failed += 1
    await state.clear()
    await message.answer(
        f"Рассылка завершена.\n\nОтправлено: {sent}\nНе доставлено: {failed}",
        reply_markup=owner_reply_kb(),
    )


@router.message(F.text == "👤 Режим пользователя")
async def owner_user_mode(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    await state.clear()
    await message.answer(
        "<b>Тестовый режим пользователя</b>\n\nВыберите роль, которую хотите проверить:",
        reply_markup=main_menu_kb(),
    )


async def main():
    if not BOT_TOKEN:
        raise RuntimeError("BOT_TOKEN is missing. Add it to Railway Variables")
    if not ADMIN_IDS:
        print("WARNING: ADMIN_ID/ADMIN_IDS is not configured; owner panel will be unavailable.")
    if not USE_POSTGRES:
        print("WARNING: DATABASE_URL is missing. SQLite fallback is NOT persistent on Railway redeploys.")
    init_db()
    bot = Bot(BOT_TOKEN, default=DefaultBotProperties(parse_mode="HTML"))
    dp = Dispatcher(storage=MemoryStorage())
    dp.include_router(router)
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
